"""
Generacion de reportes del inventario de licencias en varios formatos.

Una sola fuente de datos (`construir_dataset`) respeta los MISMOS filtros del
dashboard (tenant por ruta + empresa/tipo/origen/estado/proveedor por GET), de
modo que el reporte exportado coincide con lo que el usuario esta viendo.

Builders:
  - `excel_response`  -> .xlsx con hoja Resumen (KPIs + graficos nativos) + Detalle.
  - `csv_response`    -> .csv del detalle (compatible Excel/LibreOffice, BOM UTF-8).
El PDF se arma en la vista reutilizando el mismo dataset + una plantilla HTML.
"""
from __future__ import annotations

import csv
from datetime import timedelta
from io import BytesIO

import openpyxl
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from django.db.models import Count, Prefetch
from django.http import HttpResponse
from django.utils import timezone

from .models import Asignacion, Licencia, Tenant

NARANJA = "DF6E12"
COLUMNAS_DETALLE = [
    "Tipo Licencia", "Fabricante", "Proveedor", "Tenant", "Empresa Duena",
    "Estado", "Origen", "Factura Origen", "Fecha Inicio", "Usuario Asignado",
    "Email Usuario", "Centro de Costos", "Gerencia/Area", "Division",
    "Unidad", "Fecha Asignacion", "Fecha Vencimiento",
]


def construir_dataset(request, tenant_id=None):
    """Filtra el inventario (mismos filtros del dashboard) y calcula KPIs + series."""
    hoy = timezone.now().date()
    limite = hoy + timedelta(days=30)

    empresa_id = request.GET.get("empresa") or None
    tipo_id = request.GET.get("tipo") or None
    origen = request.GET.get("origen") or ""
    estado = request.GET.get("estado") or ""
    proveedor_id = request.GET.get("proveedor") or None

    tenant_label = None
    qs = Licencia.objects.all()
    if tenant_id:
        tenant_label = Tenant.objects.filter(pk=tenant_id).values_list("nombre", flat=True).first()
        qs = qs.filter(tenant_id=tenant_id)
    if empresa_id:
        qs = qs.filter(empresa_id=empresa_id)
    if tipo_id:
        qs = qs.filter(tipo_id=tipo_id)
    if origen:
        qs = qs.filter(origen=origen)
    if proveedor_id:
        qs = qs.filter(proveedor_id=proveedor_id)
    if estado:
        if estado == Licencia.ESTADO_ASIGNADA:
            sub = qs.filter(asignaciones__activo=True)
        elif estado == Licencia.ESTADO_VENCIDA:
            sub = qs.filter(fecha_vencimiento__lt=hoy)
        elif estado == "POR_VENCER":
            sub = qs.filter(fecha_vencimiento__gte=hoy, fecha_vencimiento__lte=limite)
        elif estado == Licencia.ESTADO_DISPONIBLE:
            sub = qs.filter(
                estado_operativo=Licencia.ESTADO_DISPONIBLE, fecha_vencimiento__gte=hoy
            ).exclude(asignaciones__activo=True)
        else:
            sub = qs.filter(estado_operativo=estado)
        qs = qs.filter(pk__in=sub.values("pk"))

    qs = qs.select_related("tipo", "empresa", "tenant", "proveedor", "factura_origen")
    # Prefetch de la asignacion activa (con el empleado y su estructura) para evitar
    # N+1 al armar el detalle.
    qs = qs.prefetch_related(Prefetch(
        "asignaciones",
        queryset=Asignacion.objects.filter(activo=True).select_related(
            "empleado__area", "empleado__division", "empleado__unidad",
        ),
        to_attr="_activas",
    ))

    kpis = {
        "total": qs.count(),
        "asignadas": qs.filter(asignaciones__activo=True).distinct().count(),
        "vencidas": qs.filter(fecha_vencimiento__lt=hoy).count(),
        "por_vencer": qs.filter(fecha_vencimiento__gte=hoy, fecha_vencimiento__lte=limite).count(),
        "disponibles": (
            qs.filter(estado_operativo=Licencia.ESTADO_DISPONIBLE, fecha_vencimiento__gte=hoy)
            .exclude(asignaciones__activo=True).distinct().count()
        ),
    }

    origen_labels = dict(Licencia.ORIGENES)
    # "Por estado" usa el estado DERIVADO (mismo criterio que la property
    # Licencia.estado), no el estado_operativo crudo, para que cuadre con los KPIs.
    _precedentes = ["SUSPENDIDA", "PENDIENTE_ACTIVACION", "REVOCADA"]
    _derivable = qs.exclude(estado_operativo__in=_precedentes)
    _no_vencida = _derivable.filter(fecha_vencimiento__gte=hoy)
    por_estado = [
        ("Disponible", _no_vencida.exclude(asignaciones__activo=True).distinct().count()),
        ("Asignada", _no_vencida.filter(asignaciones__activo=True).distinct().count()),
        ("Vencida", _derivable.filter(fecha_vencimiento__lt=hoy).count()),
        ("Suspendida", qs.filter(estado_operativo="SUSPENDIDA").count()),
        ("Pendiente de activacion", qs.filter(estado_operativo="PENDIENTE_ACTIVACION").count()),
        ("Revocada", qs.filter(estado_operativo="REVOCADA").count()),
    ]
    por_estado = [(label, n) for label, n in por_estado if n]
    por_origen = [
        (origen_labels.get(r["origen"], r["origen"]), r["total"])
        for r in qs.values("origen").annotate(total=Count("id")).order_by("-total")
    ]
    por_tipo = [
        (r["tipo__nombre"] or "Sin tipo", r["total"])
        for r in qs.values("tipo__nombre").annotate(total=Count("id")).order_by("-total")[:10]
    ]

    filtros = _describir_filtros(tenant_label, empresa_id, tipo_id, origen, estado, proveedor_id)

    return {
        "hoy": hoy,
        "limite": limite,
        "tenant_label": tenant_label,
        "licencias": qs,
        "kpis": kpis,
        "por_estado": por_estado,
        "por_origen": por_origen,
        "por_tipo": por_tipo,
        "filtros": filtros,
    }


def _describir_filtros(tenant_label, empresa_id, tipo_id, origen, estado, proveedor_id):
    partes = []
    if tenant_label:
        partes.append(f"Tenant: {tenant_label}")
    if empresa_id:
        from .models import Empresa
        nombre = Empresa.objects.filter(pk=empresa_id).values_list("nombre", flat=True).first()
        if nombre:
            partes.append(f"Empresa: {nombre}")
    if tipo_id:
        from .models import TipoLicencia
        nombre = TipoLicencia.objects.filter(pk=tipo_id).values_list("nombre", flat=True).first()
        if nombre:
            partes.append(f"Tipo: {nombre}")
    if origen:
        partes.append(f"Origen: {dict(Licencia.ORIGENES).get(origen, origen)}")
    if estado:
        label = dict(Licencia.ESTADOS_OPERATIVOS).get(estado, "Por vencer" if estado == "POR_VENCER" else estado)
        partes.append(f"Estado: {label}")
    if proveedor_id:
        from .models import Proveedor
        nombre = Proveedor.objects.filter(pk=proveedor_id).values_list("nombre", flat=True).first()
        if nombre:
            partes.append(f"Proveedor: {nombre}")
    return partes


def _fila_detalle(lic):
    activas = getattr(lic, "_activas", None)
    if activas is not None:
        asignacion = activas[0] if activas else None
    else:
        asignacion = lic.usuario_activo
    usuario_nombre, email_usuario, centro_costos = "DISPONIBLE", "-", "-"
    area_code, division_code, unidad_nombre, fecha_asig = "-", "-", "-", "-"
    if asignacion:
        emp = asignacion.empleado
        usuario_nombre = emp.nombre_completo
        email_usuario = getattr(emp, "email_principal", None) or "-"
        centro_costos = emp.centro_de_costos or "-"
        if asignacion.fecha_asignacion:
            fecha_asig = asignacion.fecha_asignacion.strftime("%d/%m/%Y")
        if emp.area:
            area_code = emp.area.codigo or "-"
        if emp.division:
            division_code = emp.division.codigo or "-"
        if getattr(emp, "unidad", None):
            unidad_nombre = emp.unidad.nombre
    return [
        lic.tipo.nombre,
        lic.tipo.fabricante,
        lic.proveedor.nombre if lic.proveedor else "Directo",
        lic.tenant.nombre,
        lic.empresa.nombre if lic.empresa else "-",
        lic.estado,
        lic.get_origen_display(),
        lic.factura_origen.numero if lic.factura_origen else "-",
        (lic.fecha_inicio or lic.fecha_compra).strftime("%d/%m/%Y") if (lic.fecha_inicio or lic.fecha_compra) else "-",
        usuario_nombre,
        email_usuario,
        centro_costos,
        area_code,
        division_code,
        unidad_nombre,
        fecha_asig,
        lic.fecha_vencimiento.strftime("%d/%m/%Y") if lic.fecha_vencimiento else "-",
    ]


# ---------------------------------------------------------------------------
# EXCEL (.xlsx) — Resumen con graficos nativos + Detalle
# ---------------------------------------------------------------------------
def excel_response(dataset):
    wb = openpyxl.Workbook()
    _hoja_resumen(wb, dataset)
    _hoja_detalle(wb, dataset)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    resp = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    fecha = timezone.now().strftime("%d-%m-%Y")
    resp["Content-Disposition"] = f'attachment; filename="Reporte_Licencias_{fecha}.xlsx"'
    return resp


def _hoja_resumen(wb, dataset):
    ws = wb.active
    ws.title = "Resumen"
    titulo = Font(bold=True, size=15, color=NARANJA)
    sub = Font(bold=True, size=11, color="FFFFFF")
    sub_fill = PatternFill(start_color=NARANJA, end_color=NARANJA, fill_type="solid")
    bold = Font(bold=True)

    ws["A1"] = "REPORTE DE LICENCIAS"
    ws["A1"].font = titulo
    ws["A2"] = "Generado: " + timezone.now().strftime("%d/%m/%Y %H:%M")
    ws["A2"].font = Font(italic=True, color="808080")
    fila = 3
    if dataset["filtros"]:
        ws.cell(row=fila, column=1, value="Filtros: " + " | ".join(dataset["filtros"]))
        ws.cell(row=fila, column=1).font = Font(italic=True, color="606060")
        fila += 1

    # --- KPIs ---
    fila += 1
    ws.cell(row=fila, column=1, value="INDICADORES").font = bold
    fila += 1
    kpis = dataset["kpis"]
    items = [
        ("Total licencias", kpis["total"]),
        ("Disponibles", kpis["disponibles"]),
        ("Asignadas", kpis["asignadas"]),
        ("Vencidas", kpis["vencidas"]),
        ("Por vencer (30 dias)", kpis["por_vencer"]),
    ]
    for label, value in items:
        ws.cell(row=fila, column=1, value=label)
        c = ws.cell(row=fila, column=2, value=value)
        c.font = bold
        fila += 1

    # --- Tablas auxiliares para los graficos (a la derecha) ---
    def tabla(col, titulo_tabla, datos):
        ws.cell(row=5, column=col, value=titulo_tabla).font = sub
        ws.cell(row=5, column=col).fill = sub_fill
        ws.cell(row=5, column=col + 1).fill = sub_fill
        r = 6
        for nombre, total in (datos or [("Sin datos", 0)]):
            ws.cell(row=r, column=col, value=str(nombre))
            ws.cell(row=r, column=col + 1, value=total)
            r += 1
        return r - 1  # ultima fila con datos

    fin_estado = tabla(4, "Por estado", dataset["por_estado"])      # cols D,E
    fin_origen = tabla(7, "Por origen", dataset["por_origen"])      # cols G,H
    fin_tipo = tabla(10, "Por tipo (Top 10)", dataset["por_tipo"])  # cols J,K

    # --- Graficos ---
    pie = PieChart()
    pie.title = "Distribucion por estado"
    pie.add_data(Reference(ws, min_col=5, min_row=6, max_row=max(fin_estado, 6)), titles_from_data=False)
    pie.set_categories(Reference(ws, min_col=4, min_row=6, max_row=max(fin_estado, 6)))
    pie.height, pie.width = 7.5, 11
    ws.add_chart(pie, "A14")

    bar_o = BarChart()
    bar_o.title = "Por origen"
    bar_o.add_data(Reference(ws, min_col=8, min_row=6, max_row=max(fin_origen, 6)), titles_from_data=False)
    bar_o.set_categories(Reference(ws, min_col=7, min_row=6, max_row=max(fin_origen, 6)))
    bar_o.legend = None
    bar_o.height, bar_o.width = 7.5, 11
    ws.add_chart(bar_o, "G14")

    bar_t = BarChart()
    bar_t.type = "bar"
    bar_t.title = "Por tipo (Top 10)"
    bar_t.add_data(Reference(ws, min_col=11, min_row=6, max_row=max(fin_tipo, 6)), titles_from_data=False)
    bar_t.set_categories(Reference(ws, min_col=10, min_row=6, max_row=max(fin_tipo, 6)))
    bar_t.legend = None
    bar_t.height, bar_t.width = 7.5, 12
    ws.add_chart(bar_t, "A31")

    for col, width in [("A", 24), ("B", 12), ("D", 16), ("E", 8), ("G", 14), ("H", 8), ("J", 22), ("K", 8)]:
        ws.column_dimensions[col].width = width


def _hoja_detalle(wb, dataset):
    ws = wb.create_sheet("Detalle")
    ws.append(COLUMNAS_DETALLE)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color=NARANJA, end_color=NARANJA, fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for lic in dataset["licencias"]:
        ws.append(_fila_detalle(lic))

    ws.freeze_panes = "A2"
    for idx in range(1, len(COLUMNAS_DETALLE) + 1):
        col = get_column_letter(idx)
        largo = max((len(str(ws.cell(row=r, column=idx).value or "")) for r in range(1, ws.max_row + 1)), default=10)
        ws.column_dimensions[col].width = min(largo + 2, 45)


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------
def csv_response(dataset):
    resp = HttpResponse(content_type="text/csv; charset=utf-8")
    fecha = timezone.now().strftime("%d-%m-%Y")
    resp["Content-Disposition"] = f'attachment; filename="Reporte_Licencias_{fecha}.csv"'
    resp.write("﻿")  # BOM: que Excel reconozca UTF-8
    writer = csv.writer(resp, delimiter=";")
    writer.writerow(COLUMNAS_DETALLE)
    for lic in dataset["licencias"]:
        writer.writerow(_fila_detalle(lic))
    return resp
