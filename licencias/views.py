import json
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill
from datetime import timedelta
import random
 
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.contrib.auth import login
from axes.utils import reset
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.contrib import messages
from django.core.cache import cache
from django.core.mail import send_mail
from django.conf import settings
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.db import transaction
from django.db.models import Count, Q
 
from empleados.models import Empleado, GerenciaDivision, GerenciaArea, Unidad
from .models import (
    Asignacion,
    Empresa,
    Licencia,
    Proveedor,
    Tenant,
    TipoLicencia,
)
from .forms import (
    EmpleadoForm,
    LicenciaForm,
    ProveedorForm,
    TipoLicenciaForm,
)
from bitacora.actions import (
    log_asignacion_licencia,
    log_baja_empleado,
    log_creacion_licencias,
    log_crear_empleado,
    log_proveedor_crear,
    log_tipo_licencia_crear,
    log_editar_licencia,
    log_editar_empleado,
    log_eliminar_licencia,
    log_eliminar_licencias_masivo,
    log_exportar_excel,
    log_liberar_licencia,
    log_reactivar_empleado,
    log_sincronizar_m365,
    log_proveedor_editar,
    log_proveedor_eliminar,
    log_tipo_licencia_editar,
    log_tipo_licencia_eliminar,
)

def exigir_permiso(request, permiso):
    if not request.user.has_perm(permiso):
        raise PermissionDenied


def exigir_algun_permiso(request, permisos):
    if not any(request.user.has_perm(permiso) for permiso in permisos):
        raise PermissionDenied

@login_required
def exportar_excel(request, tenant_id=None):
    """
    Genera un reporte consolidado en formato Excel (.xlsx) de los activos de software.
    Implementa OpenPyXL para el formateo directo del buffer de memoria sin escritura en disco.
    """
    exigir_permiso(request, 'licencias.view_licencia')
    tenant_label = None
    if tenant_id:
        tenant_label = Tenant.objects.filter(pk=tenant_id).values_list('nombre', flat=True).first()
    log_exportar_excel(request, tenant_label=tenant_label)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Completo"

    headers = [
        'Tipo Licencia', 'Fabricante', 'Proveedor', 'Tenant', 'Empresa Duena',
        'Estado', 'Origen', 'Factura Origen', 'Fecha Inicio', 'Usuario Asignado',
        'Email Usuario', 'Centro de Costos', 'Gerencia/Area', 'Division',
        'Unidad', 'Fecha Asignacion', 'Fecha Vencimiento'
    ]
    ws.append(headers)

    # Estilos de la cabecera
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="DF6E12", end_color="DF6E12", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Optimización de consultas usando select_related para evitar el problema N+1
    if tenant_id:
        licencias = Licencia.objects.filter(tenant_id=tenant_id).select_related('tipo', 'empresa', 'tenant', 'proveedor', 'factura_origen')
    else:
        licencias = Licencia.objects.all().select_related('tipo', 'empresa', 'tenant', 'proveedor', 'factura_origen')

    for lic in licencias:
        asignacion = lic.usuario_activo
        
        # Variables por defecto para registros sin asignación
        usuario_nombre = "DISPONIBLE"
        email_usuario = "-"
        centro_costos = "-"
        area_code = "-"
        division_code = "-"
        unidad_nombre = "-"
        fecha_asignacion_str = "-"

        if asignacion:
            emp = asignacion.empleado
            
            # Refresco del ORM para garantizar la captura de campos cacheados en memoria
            try:
                emp.refresh_from_db()
            except Exception:
                pass 

            usuario_nombre = emp.nombre_completo
            email_usuario = getattr(emp, 'email_principal', getattr(emp, 'email', '-')) or "-"
            centro_costos = emp.centro_de_costos if emp.centro_de_costos else "-"
            
            if asignacion.fecha_asignacion:
                fecha_asignacion_str = asignacion.fecha_asignacion.strftime('%d/%m/%Y')

            # Extracción de estructura organizacional
            if emp.area: area_code = emp.area.codigo
            if emp.division: division_code = emp.division.codigo
            if hasattr(emp, 'unidad') and emp.unidad: unidad_nombre = emp.unidad.nombre

        ws.append([
            lic.tipo.nombre,
            lic.tipo.fabricante,
            lic.proveedor.nombre if lic.proveedor else "Directo",
            lic.tenant.nombre,
            lic.empresa.nombre if lic.empresa else "-",
            lic.estado,
            lic.get_origen_display(),
            lic.factura_origen.numero if lic.factura_origen else "-",
            lic.fecha_inicio or lic.fecha_compra,
            usuario_nombre,
            email_usuario,
            centro_costos,
            area_code,
            division_code,
            unidad_nombre,
            fecha_asignacion_str, 
            lic.fecha_vencimiento,
        ])

    # Autoajuste dinámico de columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length: 
                    max_length = len(str(cell.value))
            except Exception: 
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    fecha = timezone.now().strftime('%d-%m-%Y')
    response['Content-Disposition'] = f'attachment; filename="Reporte_Licencias_{fecha}.xlsx"'
    wb.save(response)
    
    return response
# ==========================================
# MÓDULO token
# ==========================================

def validar_token_bloqueo(request):
    """
    Pantalla de desbloqueo tras intentos fallidos.

    Token y dedup viven en `request.session` (no en cache global) para
    funcionar consistente entre workers WSGI. El envio automatico se dispara
    en GET cuando el signal `preparar_desbloqueo` dejo el flag
    `enviar_token_pendiente` en sesion.
    """
    from .services import (
        enviar_token_desbloqueo,
        validar_token_desbloqueo,
        limpiar_token_desbloqueo,
    )

    username = request.session.get('usuario_bloqueado_nombre')
    if not username:
        return redirect('login')

    # --- Auto-envio diferido desde el signal user_locked_out ---
    if request.method == 'GET' and request.session.get('enviar_token_pendiente'):
        user = User.objects.filter(username=username).first()
        if user:
            ok, info = enviar_token_desbloqueo(
                request=request, user=user, source="autosend"
            )
            if ok:
                request.session.pop('enviar_token_pendiente', None)
                request.session.modified = True
                messages.info(request, info)
            else:
                messages.warning(
                    request,
                    f"{info} Pulsa 'Reenviar codigo' para intentarlo nuevamente.",
                )

    # --- Validacion de codigo ingresado ---
    if request.method == 'POST':
        token_ingresado = (request.POST.get('token') or '').strip()

        if validar_token_desbloqueo(request=request, token_ingresado=token_ingresado):
            user = User.objects.filter(username=username).first()
            if not user:
                messages.error(request, "Usuario no encontrado.")
                return redirect('login')

            # Limpia el bloqueo de Axes (username + ip).
            try:
                reset(username=username, ip=request.META.get('REMOTE_ADDR'))
            except Exception:
                # No interrumpir el login si la limpieza de axes falla.
                pass

            login(request, user, backend='django.contrib.auth.backends.ModelBackend')

            limpiar_token_desbloqueo(request)
            request.session.pop('usuario_bloqueado_nombre', None)
            request.session.pop('enviar_token_pendiente', None)
            request.session.modified = True

            messages.success(request, "Acceso concedido mediante codigo de seguridad.")
            return redirect('dashboard_general')

        messages.error(request, "Codigo incorrecto o expirado.")

    return render(request, 'registration/desbloqueo_token.html')


def enviar_token_bloqueo(request):
    """
    Reenvio manual de token desde la pantalla /desbloqueo-seguro/.
    Solo POST. Delega al service.
    """
    from .services import enviar_token_desbloqueo

    username = request.session.get('usuario_bloqueado_nombre')
    if not username:
        return redirect('login')

    if request.method != 'POST':
        return redirect('validar_token_bloqueo')

    user = User.objects.filter(username=username).first()
    if not user:
        messages.error(request, "No se pudo enviar el codigo: usuario no encontrado.")
        return redirect('validar_token_bloqueo')

    ok, info = enviar_token_desbloqueo(request=request, user=user, source="manual")
    if ok:
        messages.success(request, info)
    else:
        messages.warning(request, info)
    return redirect('validar_token_bloqueo')
# ==========================================
# MÓDULO PRINCIPAL (DASHBOARD Y KPIS)
# ==========================================

@login_required
def inicio(request):
    """Pantalla de entrada: accesos y acciones principales, sin reporteria."""
    context = {
        'titulo': 'Inicio',
        'fecha_actual': timezone.now(),
    }
    return render(request, 'licencias/inicio.html', context)


@login_required
def dashboard(request, tenant_id=None):
    """Dashboard ejecutivo: reporte general de inventario, uso y riesgos."""
    exigir_permiso(request, 'licencias.view_licencia')

    tenants = Tenant.objects.filter(activo=True).order_by('nombre')
    hoy = timezone.now().date()
    limite_30_dias = hoy + timedelta(days=30)
    empresa_id = request.GET.get('empresa') or None
    tipo_id = request.GET.get('tipo') or None
    origen = request.GET.get('origen') or ''

    if tenant_id:
        tenant_seleccionado = get_object_or_404(Tenant, pk=tenant_id)
        licencias = Licencia.objects.filter(tenant_id=tenant_id)
        titulo = f"Dashboard de {tenant_seleccionado.nombre}"
    else:
        tenant_seleccionado = None
        licencias = Licencia.objects.all()
        titulo = "Dashboard general"

    if empresa_id:
        licencias = licencias.filter(empresa_id=empresa_id)
    if tipo_id:
        licencias = licencias.filter(tipo_id=tipo_id)
    if origen:
        licencias = licencias.filter(origen=origen)

    licencias = licencias.select_related('tipo', 'empresa', 'tenant', 'proveedor', 'factura_origen')
    total_licencias = licencias.count()
    asignadas = licencias.filter(asignaciones__activo=True).distinct().count()
    vencidas = licencias.filter(fecha_vencimiento__lt=hoy).count()
    por_vencer = licencias.filter(fecha_vencimiento__gte=hoy, fecha_vencimiento__lte=limite_30_dias).count()
    disponibles = (
        licencias
        .filter(estado_operativo=Licencia.ESTADO_DISPONIBLE, fecha_vencimiento__gte=hoy)
        .exclude(asignaciones__activo=True)
        .distinct()
        .count()
    )

    stock_bajo = [
        tipo for tipo in TipoLicencia.objects.filter(activo=True).select_related('proveedor_default')
        if tipo.stock_minimo and tipo.cantidad_disponible <= tipo.stock_minimo
    ][:8]

    def chart_rows(rows):
        values = [{'label': label, 'value': int(value or 0)} for label, value in rows]
        max_value = max([row['value'] for row in values] or [1])
        for row in values:
            row['pct'] = 0 if max_value == 0 else round(row['value'] * 100 / max_value, 2)
        return values

    estado_labels = dict(Licencia.ESTADOS_OPERATIVOS)
    origen_labels = dict(Licencia.ORIGENES)
    estado_rows = chart_rows(
        (estado_labels.get(row['estado_operativo'], row['estado_operativo']), row['total'])
        for row in licencias.values('estado_operativo').annotate(total=Count('id')).order_by('-total')
    )
    origen_rows = chart_rows(
        (origen_labels.get(row['origen'], row['origen']), row['total'])
        for row in licencias.values('origen').annotate(total=Count('id')).order_by('-total')
    )
    tipo_rows = chart_rows(
        (row['tipo__nombre'] or 'Sin tipo', row['total'])
        for row in licencias.values('tipo__nombre').annotate(total=Count('id')).order_by('-total')[:8]
    )

    context = {
        'titulo': titulo,
        'tenants': tenants,
        'tenant_seleccionado': tenant_seleccionado,
        'empresa_filtro': empresa_id,
        'tipo_filtro': tipo_id,
        'origen_filtro': origen,
        'empresas': Empresa.objects.filter(activo=True).select_related('tenant').order_by('tenant__nombre', 'nombre'),
        'tipos_licencia': TipoLicencia.objects.filter(activo=True).order_by('fabricante', 'nombre'),
        'origenes_licencia': Licencia.ORIGENES,
        'chart_estado_rows': estado_rows,
        'chart_origen_rows': origen_rows,
        'chart_tipo_rows': tipo_rows,
        'kpi_total': total_licencias,
        'kpi_ocupadas': asignadas,
        'kpi_disponibles': disponibles,
        'kpi_vencidas': vencidas,
        'kpi_por_vencer': por_vencer,
        'kpi_empresas': Empresa.objects.filter(activo=True).count(),
        'kpi_empleados': Empleado.objects.filter(activo=True).count(),
        'kpi_proveedores': Proveedor.objects.filter(activo=True).count(),
        'stock_bajo': stock_bajo,
        'ultimas_asignaciones': (
            Asignacion.objects
            .filter(activo=True, licencia__in=licencias)
            .select_related('licencia__tipo', 'licencia__empresa', 'empleado')
            .order_by('-fecha_asignacion')[:6]
        ),
        'licencias_por_vencer': licencias.filter(fecha_vencimiento__gte=hoy, fecha_vencimiento__lte=limite_30_dias).order_by('fecha_vencimiento')[:8],
    }
    return render(request, 'licencias/dashboard_reportes.html', context)


@login_required
def gestionar_licencias(request, tenant_id=None):
    """
    Controlador principal de la vista gerencial. 
    Calcula KPIs operativos (Asignaciones, Disponibilidad, Riesgo de Vencimiento)
    mediante iteración única para minimizar la carga transaccional en la base de datos.
    """
    exigir_permiso(request, 'licencias.view_licencia')
    from .application.use_cases import uc_listar_licencias

    tenants = Tenant.objects.all()
    q = request.GET.get('q', '').strip()
    estado = request.GET.get('estado', '').strip()
    empresa_id = None
    proveedor_id = request.GET.get('proveedor') or None
    tipo_id = request.GET.get('tipo') or None
    precio_min = request.GET.get('precio_min') or ''
    precio_max = request.GET.get('precio_max') or ''
    fecha_desde = request.GET.get('fecha_desde') or None
    fecha_hasta = request.GET.get('fecha_hasta') or None

    tenant_filtro = None

    if tenant_filtro:
        tenant_seleccionado = get_object_or_404(Tenant, pk=tenant_filtro)
        licencias = uc_listar_licencias(
            tenant_id=tenant_filtro,
            empresa_id=empresa_id,
            proveedor_id=proveedor_id,
            tipo_id=tipo_id,
            estado=estado,
            q=q,
            fecha_desde=fecha_desde,
            fecha_hasta=fecha_hasta,
        )
        titulo = f"Licencias de {tenant_seleccionado.nombre}"
    else:
        tenant_seleccionado = None
        licencias = uc_listar_licencias(
            empresa_id=empresa_id,
            proveedor_id=proveedor_id,
            tipo_id=tipo_id,
            estado=estado,
            q=q,
            fecha_desde=fecha_desde,
            fecha_hasta=fecha_hasta,
        )
        titulo = "Todas las Licencias"

    if precio_min:
        licencias = licencias.filter(tipo__precio_venta__gte=precio_min)
    if precio_max:
        licencias = licencias.filter(tipo__precio_venta__lte=precio_max)

    empleados = Empleado.objects.filter(activo=True).select_related('empresa', 'area').order_by('nombre_completo')

    hoy = timezone.now().date()
    limite_30_dias = hoy + timedelta(days=30)

    # Cálculo algorítmico de KPIs
    total_licencias = licencias.count()
    ocupadas = 0
    disponibles = 0
    vencidas = 0
    por_vencer = 0

    for lic in licencias:
        if lic.usuario_activo:
            ocupadas += 1
            
        # Evaluación de riesgos y caducidad
        if lic.fecha_vencimiento < hoy:
            vencidas += 1
        elif lic.fecha_vencimiento <= limite_30_dias:
            por_vencer += 1
            
        # Disponibilidad neta
        if lic.puede_asignarse:
            disponibles += 1

    form_licencia = LicenciaForm()
    catalogo_base = Licencia.objects.all()
    if tenant_filtro:
        catalogo_base = catalogo_base.filter(tenant_id=tenant_filtro)
    if empresa_id:
        catalogo_base = catalogo_base.filter(empresa_id=empresa_id)
    if proveedor_id:
        catalogo_base = catalogo_base.filter(proveedor_id=proveedor_id)
    if tipo_id:
        catalogo_base = catalogo_base.filter(tipo_id=tipo_id)

    proveedores_filtrados = Proveedor.objects.filter(activo=True)
    tipos_filtrados = TipoLicencia.objects.filter(activo=True)
    if tenant_filtro or empresa_id:
        proveedores_ids = catalogo_base.exclude(proveedor_id__isnull=True).values('proveedor_id')
        tipos_ids = catalogo_base.values('tipo_id')
        proveedores_filtrados = proveedores_filtrados.filter(id__in=proveedores_ids)
        tipos_filtrados = tipos_filtrados.filter(id__in=tipos_ids)
    
    context = {
        'tenants': tenants,
        'tenant_seleccionado': tenant_seleccionado,
        'licencias': licencias,
        'empleados': empleados,
        'titulo': titulo,
        
        # Métricas de telemetría
        'kpi_total': total_licencias,
        'kpi_ocupadas': ocupadas,
        'kpi_disponibles': disponibles,
        'kpi_vencidas': vencidas,
        'kpi_por_vencer': por_vencer,
        
        'hoy': hoy,
        'limite_30_dias': limite_30_dias,
        'form_licencia': form_licencia,
        'q': q,
        'tenant_filtro': str(tenant_filtro) if tenant_filtro else '',
        'estado_filtro': estado,
        'empresa_filtro': empresa_id,
        'proveedor_filtro': proveedor_id,
        'tipo_filtro': tipo_id,
        'precio_min': precio_min,
        'precio_max': precio_max,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'empresas': Empresa.objects.filter(activo=True).select_related('tenant').order_by('tenant__nombre', 'nombre'),
        'proveedores': proveedores_filtrados.order_by('nombre'),
        'tipos_licencia': tipos_filtrados.order_by('fabricante', 'nombre'),
        'estados_licencia': Licencia.ESTADOS_OPERATIVOS,
        'licencias_active': 'dashboard',
    }
    return render(request, 'licencias/dashboard.html', context)


# ==========================================
# MÓDULO DE TRANSACCIONES DE LICENCIAS
# ==========================================

@login_required
def asignaciones_licencias(request):
    """Gestion operativa de asignaciones: stock disponible + asignaciones activas."""
    exigir_permiso(request, 'licencias.view_licencia')
    from .application.use_cases import uc_listar_licencias

    tenant_id = request.GET.get('tenant') or None
    empresa_id = request.GET.get('empresa') or None
    tipo_id = request.GET.get('tipo') or None
    empleado_id = request.GET.get('empleado') or None
    estado = request.GET.get('estado') or 'activas'

    licencias = uc_listar_licencias(
        tenant_id=tenant_id,
        empresa_id=empresa_id,
        tipo_id=tipo_id,
        estado=Licencia.ESTADO_DISPONIBLE,
    )
    licencias = licencias.order_by('-factura_origen__fecha', 'tipo__nombre', 'fecha_vencimiento')
    tipos_disponibles = TipoLicencia.objects.filter(id__in=licencias.values('tipo_id')).order_by('fabricante', 'nombre')
    asignaciones = (
        Asignacion.objects
        .select_related('licencia__tipo', 'licencia__tenant', 'licencia__empresa', 'empleado__empresa')
        .order_by('licencia__empresa__nombre', 'empleado__nombre_completo')
    )
    if estado == 'activas':
        asignaciones = asignaciones.filter(activo=True)
    elif estado == 'liberadas':
        asignaciones = asignaciones.filter(activo=False)

    if tenant_id:
        asignaciones = asignaciones.filter(licencia__tenant_id=tenant_id)
    if empresa_id:
        asignaciones = asignaciones.filter(licencia__empresa_id=empresa_id)
    if tipo_id:
        asignaciones = asignaciones.filter(licencia__tipo_id=tipo_id)
    if empleado_id:
        asignaciones = asignaciones.filter(empleado_id=empleado_id)

    empleados = Empleado.objects.filter(activo=True).select_related('empresa', 'area').order_by('nombre_completo')
    if tenant_id:
        empleados = empleados.filter(empresa__tenant_id=tenant_id)
    if empresa_id:
        empleados = empleados.filter(empresa_id=empresa_id)

    licencias_kpi = Licencia.objects.all()
    asignaciones_kpi = Asignacion.objects.select_related('licencia')
    if tenant_id:
        licencias_kpi = licencias_kpi.filter(tenant_id=tenant_id)
        asignaciones_kpi = asignaciones_kpi.filter(licencia__tenant_id=tenant_id)
    if empresa_id:
        licencias_kpi = licencias_kpi.filter(empresa_id=empresa_id)
        asignaciones_kpi = asignaciones_kpi.filter(licencia__empresa_id=empresa_id)
    if tipo_id:
        licencias_kpi = licencias_kpi.filter(tipo_id=tipo_id)
        asignaciones_kpi = asignaciones_kpi.filter(licencia__tipo_id=tipo_id)
    if empleado_id:
        asignaciones_kpi = asignaciones_kpi.filter(empleado_id=empleado_id)

    # KPIs del modulo sobre el contexto filtrado.
    from datetime import timedelta
    hoy = timezone.now().date()
    prox_30 = hoy + timedelta(days=30)
    kpi = {
        'activas':     asignaciones_kpi.filter(activo=True).count(),
        'liberadas':   asignaciones_kpi.filter(activo=False).count(),
        'disponibles': licencias_kpi.filter(fecha_vencimiento__gte=hoy).exclude(asignaciones__activo=True).count(),
        'prox_vencer': licencias_kpi.filter(fecha_vencimiento__gte=hoy, fecha_vencimiento__lte=prox_30).count(),
        'vencidas':    licencias_kpi.filter(fecha_vencimiento__lt=hoy).count(),
    }

    context = {
        'licencias': licencias,
        'asignaciones': asignaciones,
        'empleados': empleados,
        'tenants': Tenant.objects.filter(activo=True).order_by('nombre'),
        'empresas': Empresa.objects.filter(activo=True).select_related('tenant').order_by('tenant__nombre', 'nombre'),
        'tipos_licencia': TipoLicencia.objects.filter(activo=True).order_by('fabricante', 'nombre'),
        'tipos_disponibles': tipos_disponibles,
        'puede_asignar': request.user.is_superuser or request.user.has_perm('licencias.add_asignacion'),
        'filtros': {
            'tenant_id': tenant_id,
            'empresa_id': empresa_id,
            'tipo_id': tipo_id,
            'empleado_id': empleado_id,
        },
        'kpi': kpi,
        'estado': estado,
        'licencias_active': 'gestionar',
    }
    return render(request, 'licencias/asignaciones.html', context)


@login_required
def asignar_licencia(request, licencia_id):
    """Vincula un activo a un empleado. Delega a uc_asignar_licencia."""
    from .application.use_cases import uc_asignar_licencia
    from .infrastructure import repositories as repo
    exigir_permiso(request, 'licencias.add_asignacion')
    if request.method == 'POST':
        licencia = repo.get_licencia(licencia_id)
        empleado_id = request.POST.get('empleado_id')
        empleado = get_object_or_404(Empleado.objects.select_related('empresa'), pk=empleado_id)

        ok, info = uc_asignar_licencia(request=request, licencia=licencia, empleado=empleado)
        if ok:
            messages.success(request, info)
        else:
            (messages.error if 'concurrencia' in info.lower() else messages.warning)(request, info)

    return redirect(request.META.get('HTTP_REFERER', 'gestionar_licencias'))


@login_required
def liberar_licencia(request, licencia_id):
    """Revoca el acceso a un activo. Delega a uc_liberar_licencia."""
    from .application.use_cases import uc_liberar_licencia
    from .infrastructure import repositories as repo
    exigir_permiso(request, 'licencias.change_asignacion')
    licencia = repo.get_licencia(licencia_id)

    ok, info, _ = uc_liberar_licencia(request=request, licencia=licencia)
    if ok:
        messages.success(request, info)
    else:
        messages.warning(request, info)

    return redirect('gestionar_licencias')


# ==========================================
# MÓDULO DE GESTIÓN DE IDENTIDADES (EMPLEADOS)
# ==========================================

@login_required
def lista_empleados(request):
    """Gestor principal del directorio de identidades y altas de personal."""
    if request.method == 'POST':
        exigir_permiso(request, 'empleados.add_empleado')
    else:
        exigir_permiso(request, 'empleados.view_empleado')
    empleados = Empleado.objects.all().select_related('empresa', 'area', 'division').order_by('nombre_completo')
    
    if request.method == 'POST':
        form = EmpleadoForm(request.POST)
        if form.is_valid():
            empleado_creado = form.save()
            log_crear_empleado(request, empleado_creado)
            messages.success(request, "Registro de identidad completado con éxito.")
            return redirect('lista_empleados')
        else:
            messages.error(request, "Fallo de validación: Verifique conflictos de CI o Correo Electrónico.")
    else:
        form = EmpleadoForm()

    context = {
        'empleados': empleados,
        'form': form,
        'titulo': 'Directorio de Identidades'
    }
    return render(request, 'empleados.html', context)


@login_required
def editar_empleado(request, empleado_id):
    """Actualización de metadata operativa de una identidad existente."""
    exigir_permiso(request, 'empleados.change_empleado')
    empleado = get_object_or_404(Empleado, id=empleado_id)
    
    if request.method == 'POST':
        form = EmpleadoForm(request.POST, instance=empleado)
        if form.is_valid():
            empleado_actualizado = form.save()
            log_editar_empleado(request, empleado_actualizado)
            messages.success(request, f"Metadata de {empleado.nombre_completo} actualizada correctamente.")
            return redirect('lista_empleados')
    else:
        form = EmpleadoForm(instance=empleado)

    context = {
        'form': form,
        'empleado': empleado,
        'titulo': f'Editar Identidad: {empleado.nombre_completo}'
    }
    return render(request, 'editar_empleado.html', context)
# ==========================================
# MÓDULO DE CICLO DE VIDA DEL EMPLEADO
# ==========================================

@login_required
def baja_empleado(request, empleado_id):
    """
    Inhabilita operativamente a un colaborador y ejecuta la política de
    revocación automática de todos sus activos de software vinculados.
    """
    exigir_permiso(request, 'empleados.change_empleado')
    empleado = get_object_or_404(Empleado, id=empleado_id)
    
    # 1. Inhabilitación de la identidad
    empleado.activo = False
    empleado.save()
    
    # Importación local para mitigación de dependencias circulares
    from .models import Asignacion
    asignaciones_activas = Asignacion.objects.filter(empleado=empleado, activo=True).select_related('licencia')
    
    # 2. Revocación en lote de activos asignados
    licencias_liberadas = 0
    for asignacion in asignaciones_activas:
        licencia = asignacion.licencia
        asignacion.estado = 'LIBERADA'
        asignacion.activo = False
        asignacion.fecha_retiro = timezone.now()
        asignacion.observaciones = f"Revocación automatizada por baja operativa el {timezone.now().strftime('%d/%m/%Y')}."
        asignacion.save()
        licencia.estado_operativo = Licencia.ESTADO_VENCIDA if licencia.esta_vencida else Licencia.ESTADO_DISPONIBLE
        licencia.save(update_fields=['estado_operativo'])
        licencias_liberadas += 1
        
    mensaje = f"Baja operativa procesada para {empleado.nombre_completo}."
    if licencias_liberadas > 0:
        mensaje += f" Se liberaron {licencias_liberadas} activos vinculados."
        
    messages.success(request, mensaje)
    log_baja_empleado(request, empleado, licencias_liberadas=licencias_liberadas)
    return redirect('lista_empleados')


@login_required
def reactivar_empleado(request, empleado_id):
    """Restablece el estado operativo de una identidad previamente inhabilitada."""
    exigir_permiso(request, 'empleados.change_empleado')
    empleado = get_object_or_404(Empleado, id=empleado_id)
    empleado.activo = True
    empleado.save()
    
    log_reactivar_empleado(request, empleado)
    messages.success(request, f"Identidad operativa restablecida: {empleado.nombre_completo}.")
    return redirect('lista_empleados')

# ==========================================
# ENDPOINTS ASÍNCRONOS (AJAX / CASCADAS)
# ==========================================

@login_required
def cargar_unidades(request):
    """Endpoint API para listado dinámico de Unidades filtradas por Área."""
    exigir_permiso(request, 'empleados.view_unidad')
    area_id = request.GET.get('area_id')
    unidades = Unidad.objects.filter(area_id=area_id).order_by('nombre') if area_id else Unidad.objects.none()
    return JsonResponse(list(unidades.values('id', 'nombre')), safe=False)


@login_required
def cargar_areas(request):
    """Endpoint API para listado dinámico de Áreas formateadas para Select2."""
    exigir_permiso(request, 'empleados.view_gerenciaarea')
    empresa_id = request.GET.get('empresa_id')
    areas_list = []
    
    if empresa_id:
        areas = GerenciaArea.objects.filter(empresa_id=empresa_id).order_by('nombre')
        for area in areas:
            codigo = area.codigo if area.codigo else "S/C"
            areas_list.append({'id': area.id, 'texto': f"{codigo} - {area.nombre}"})
            
    return JsonResponse(areas_list, safe=False)


@login_required
def cargar_divisiones(request):
    """Endpoint API para listado dinámico de Divisiones filtradas por Empresa."""
    exigir_permiso(request, 'empleados.view_gerenciadivision')
    empresa_id = request.GET.get('empresa_id')
    divisiones_list = []
    
    if empresa_id:
        divisiones = GerenciaDivision.objects.filter(empresa_id=empresa_id).order_by('codigo')
        for div in divisiones:
            divisiones_list.append({'id': div.id, 'texto': f"{div.codigo} - {div.nombre}"})
            
    return JsonResponse(divisiones_list, safe=False)


@login_required
def cargar_empresas(request):
    """Endpoint API para despliegue en cascada: Tenant -> Empresa."""
    exigir_permiso(request, 'licencias.view_empresa')
    tenant_id = request.GET.get('tenant_id')
    empresas = Empresa.objects.filter(tenant_id=tenant_id).order_by('nombre') if tenant_id else Empresa.objects.none()
    return JsonResponse(list(empresas.values('id', 'nombre')), safe=False)


@login_required
def cargar_empleados(request):
    """Endpoint API para cascada Empresa -> Empleado."""
    exigir_permiso(request, 'empleados.view_empleado')
    empresa_id = request.GET.get('empresa_id')
    empleados = Empleado.objects.filter(activo=True).select_related('empresa').order_by('nombre_completo')
    if empresa_id:
        empleados = empleados.filter(empresa_id=empresa_id)
    else:
        empleados = Empleado.objects.none()

    data = [
        {
            'id': empleado.id,
            'nombre': empleado.nombre_completo,
            'email': empleado.email_principal,
            'empresa': empleado.empresa.nombre if empleado.empresa_id else '',
        }
        for empleado in empleados
    ]
    return JsonResponse(data, safe=False)


# ==========================================
# MÓDULO DE CONFIGURACIÓN GLOBAL UNIFICADO
# ==========================================



@login_required
def editar_licencia(request, licencia_id):
    """Actualizacion de atributos de un activo de software. Delega a uc_editar_licencia."""
    from .application.use_cases import uc_editar_licencia
    from .infrastructure import repositories as repo
    exigir_permiso(request, 'licencias.change_licencia')
    licencia = repo.get_licencia(licencia_id)

    if request.method == 'POST':
        form = LicenciaForm(request.POST, instance=licencia)
        if form.is_valid():
            uc_editar_licencia(request=request, form=form, licencia=licencia)
            messages.success(request, f"Parametros del activo {licencia.tipo.nombre} actualizados.")
            return redirect('gestionar_licencias')
        else:
            messages.error(request, "Error de validacion en la actualizacion del activo.")
    else:
        form = LicenciaForm(instance=licencia)

    context = {
        'form': form,
        'licencia': licencia,
        'titulo': f'Edicion de Activo: {licencia.tipo.nombre}'
    }
    return render(request, 'licencias/editar_licencia.html', context)


@login_required
def editar_proveedor(request, pk):
    """Modificacion de parametros de un socio comercial. Delega a uc_editar_proveedor."""
    from .application.use_cases import uc_editar_proveedor
    from .infrastructure import repositories as repo
    exigir_permiso(request, 'licencias.change_proveedor')
    proveedor = repo.get_proveedor(pk)
    if request.method == 'POST':
        form = ProveedorForm(request.POST, instance=proveedor)
        if form.is_valid():
            uc_editar_proveedor(request=request, form=form, proveedor=proveedor)
            messages.success(request, "Datos de proveedor comercial actualizados.")
            return redirect('catalogo_licencias')
    else:
        form = ProveedorForm(instance=proveedor)
    return render(request, 'licencias/editar_catalogo.html', {'form': form, 'titulo': f'Editar Proveedor: {proveedor.nombre}', 'active_tab': 'proveedores'})


@login_required
def editar_tipo_licencia(request, pk):
    """Edicion del catalogo de SKUs de software. Delega a uc_editar_tipo_licencia."""
    from .application.use_cases import uc_editar_tipo_licencia
    from .infrastructure import repositories as repo
    exigir_permiso(request, 'licencias.change_tipolicencia')
    tipo = repo.get_tipo_licencia(pk)
    if request.method == 'POST':
        form = TipoLicenciaForm(request.POST, instance=tipo)
        if form.is_valid():
            uc_editar_tipo_licencia(request=request, form=form, tipo=tipo)
            messages.success(request, "Especificaciones de SKU de software actualizadas.")
            return redirect('gestionar_licencias')
    else:
        form = TipoLicenciaForm(instance=tipo)
    return render(request, 'licencias/editar_catalogo.html', {'form': form, 'titulo': f'Editar Licencia: {tipo.nombre}', 'active_tab': 'licencias'})


# ==========================================
# MÓDULO DE ELIMINACIÓN Y PURGA DE DATOS
# ==========================================

@login_required
def eliminar_proveedor(request, pk):
    """Eliminacion fisica de proveedor. Delega a uc_eliminar_proveedor."""
    from .application.use_cases import uc_eliminar_proveedor
    from .infrastructure import repositories as repo
    exigir_permiso(request, 'licencias.delete_proveedor')
    proveedor = repo.get_proveedor(pk)
    uc_eliminar_proveedor(request=request, proveedor=proveedor)
    messages.success(request, "Proveedor retirado del catalogo.")
    return redirect('catalogo_licencias')

@login_required
def eliminar_tipo_licencia(request, pk):
    """Eliminacion fisica de tipo de licencia. Delega a uc_eliminar_tipo_licencia."""
    from .application.use_cases import uc_eliminar_tipo_licencia
    from .infrastructure import repositories as repo
    exigir_permiso(request, 'licencias.delete_tipolicencia')
    tipo = repo.get_tipo_licencia(pk)
    uc_eliminar_tipo_licencia(request=request, tipo=tipo)
    messages.success(request, "SKU de software retirado del catalogo global.")
    return redirect('gestionar_licencias')

@login_required
def eliminar_licencia(request, licencia_id):
    """Destruccion fisica del registro de licencia. Delega a uc_eliminar_licencia."""
    from .application.use_cases import uc_eliminar_licencia
    from .infrastructure import repositories as repo
    exigir_permiso(request, 'licencias.delete_licencia')
    licencia = repo.get_licencia(licencia_id)
    nombre_licencia = licencia.tipo.nombre

    ok, info = uc_eliminar_licencia(request=request, licencia=licencia)
    if ok:
        messages.success(request, f"El activo {nombre_licencia} ha sido eliminado permanentemente del inventario.")
    else:
        messages.error(request, info)
    return redirect('gestionar_licencias')


# ==========================================
# MÓDULO DE APROVISIONAMIENTO DE INVENTARIO
# ==========================================

@login_required
def crear_licencia(request):
    """
    Controlador para la ingesta manual de nuevos activos de software.
    Implementa inserción masiva (Bulk Create) para optimizar transacciones I/O
    y minimizar la latencia en la base de datos durante cargas de alto volumen.
    """
    exigir_permiso(request, 'licencias.add_licencia')
    if request.method == 'POST':
        form = LicenciaForm(request.POST)
        
        # 1. Captura del volumen de instancias a aprovisionar
        cantidad = int(request.POST.get('cantidad_masiva', 1))
        
        if form.is_valid():
            licencia_base = form.save(commit=False)
            nuevas_licencias = []
            
            # 2. Generación del lote de activos en memoria RAM
            for _ in range(cantidad):
                nueva_lic = Licencia(
                    tipo=licencia_base.tipo,
                    empresa=licencia_base.empresa,
                    tenant=licencia_base.tenant,
                    proveedor=licencia_base.proveedor,
                    estado_operativo=licencia_base.estado_operativo,
                    origen=Licencia.ORIGEN_MANUAL,
                    fecha_compra=licencia_base.fecha_compra,
                    fecha_inicio=licencia_base.fecha_inicio,
                    fecha_activacion=licencia_base.fecha_activacion,
                    fecha_vencimiento=licencia_base.fecha_vencimiento,
                    observaciones=licencia_base.observaciones,
                )
                nuevas_licencias.append(nueva_lic)
            
            # 3. Ejecución de Bulk Create (Transacción optimizada)
            Licencia.objects.bulk_create(nuevas_licencias)
            log_creacion_licencias(request, licencia_base, cantidad=cantidad)
            
            if cantidad > 1:
                messages.success(request, f"Aprovisionamiento masivo exitoso: {cantidad} instancias de '{licencia_base.tipo.nombre}' registradas.")
            else:
                messages.success(request, "Activo de software ingresado al inventario exitosamente.")
                
            return redirect('gestionar_licencias')
            
        else:
            messages.error(request, "Fallo de validación estructural. Verifique los parámetros de entrada.")
            
    return redirect('gestionar_licencias')


# ==========================================
# MOTOR DE SINCRONIZACIÓN AUTOMATIZADA M365
# ==========================================

@login_required
def sincronizar_m365(request):
    """
    Motor de ingesta y conciliación de datos contra el proveedor M365.
    Procesa múltiples hojas de cálculo vía DataFrames y ejecuta actualización
    transaccional garantizando el principio ACID en la base de datos.
    """
    exigir_permiso(request, 'licencias.change_licencia')
    if request.method == 'POST' and request.FILES.get('archivo_excel'):
        excel_file = request.FILES['archivo_excel']
        
        try:
            # 1. Extracción de datos en memoria mediante motor OpenPyXL
            diccionario_hojas = pd.read_excel(excel_file, engine='openpyxl', sheet_name=None)
            df = pd.concat(diccionario_hojas.values(), ignore_index=True)
            df.columns = df.columns.str.strip()

            creados, actualizados, asignadas, liberadas, sin_stock = 0, 0, 0, 0, 0

            # 2. Diccionario de enrutamiento de dominios corporativos
            mapa_dominios = {
                'avicolasofia.com': 'AVICOLASOFIA',
                'mamayatech.com': 'MAMAYATECH',
                'indatta.com': 'INDATTA',
                'agrosofiaservicios.com': 'AGROSOFIA'
            }

            # 3. Diccionario de homologación de SKUs (M365 vs Sistema Interno)
            mapa_licencias = {
                'ENTERPRISEPACK': 'E3',                             
                'SPE_F1': 'F3',                                     
                'O365_BUSINESS_PREMIUM': 'ESTANDAR',                
                'Office_365_E3_(no_Teams)': 'E3 (SIN TEAMS)',       
                'Office_365_E1_(no_Teams)': 'E1 (SIN TEAMS)',       
                'Microsoft_Teams_Enterprise_New': 'TEAMS ENTERPRISE',
                'PROJECTPROFESSIONAL': 'PROJECT',                   
                'POWER_BI_PRO': 'POWER BI',                         
                'VISIOCLIENT': 'VISIO PLAN 2',                      
                'EXCHANGESTANDARD': 'EXCHANGE ONLINE (PLAN 1)'      
            }

            # 4. Bloque Transaccional Atómico (Evita inconsistencias ante errores críticos)
            with transaction.atomic():
                for index, row in df.iterrows():
                    correo = str(row.get('Nombre principal de usuario', '')).strip().lower()
                    nombre = str(row.get('Nombre para mostrar', '')).strip()
                    tiene_lic = str(row.get('Tiene licencia', '')).strip().upper() == 'VERDADERO'
                    skus_str = str(row.get('AssignedProductSkus', '')).strip()

                    # Limpieza de registros inválidos
                    if pd.isna(correo) or not correo or '@' not in correo:
                        continue

                    # Resolución de dependencias corporativas
                    dominio = correo.split('@')[1]
                    nombre_empresa = mapa_dominios.get(dominio, 'MAMAYATECH')
                    empresa_obj = Empresa.objects.filter(nombre__icontains=nombre_empresa).first()
                    
                    if not empresa_obj: 
                        continue
                        
                    area_por_defecto = GerenciaArea.objects.filter(empresa=empresa_obj).first()

                    # Consolidación de Identidades Operativas
                    empleado, creado = Empleado.objects.get_or_create(
                        email_principal=correo,
                        defaults={
                            'nombre_completo': nombre,
                            'ci': correo.split('@')[0],
                            'empresa': empresa_obj,
                            'area': area_por_defecto,
                            'activo': tiene_lic
                        }
                    )

                    if creado: 
                        creados += 1
                    else:
                        if empleado.activo != tiene_lic:
                            empleado.activo = tiene_lic
                            empleado.save(update_fields=['activo'])
                        actualizados += 1

                    # Parseo estructural de SKUs compuestos
                    skus_lista = skus_str.split('+') if skus_str else []
                    skus_principales = []
                    skus_extra = [] 

                    for sku in skus_lista:
                        sku = sku.strip()
                        if sku in mapa_licencias:
                            skus_principales.append(sku)
                        elif sku:
                            skus_extra.append(sku)

                    nombres_licencias_excel = [mapa_licencias[s] for s in skus_principales]
                    nombres_manejados = list(mapa_licencias.values())

                    # Políticas de retención: Se revocan únicamente las licencias gestionadas en la conciliación actual
                    asignaciones_actuales = Asignacion.objects.filter(empleado=empleado, activo=True).select_related('licencia__tipo')
                    
                    for asig in asignaciones_actuales:
                        nombre_bd = asig.licencia.tipo.nombre.upper()
                        es_gestionada = any(nm.upper() in nombre_bd for nm in nombres_manejados)
                        
                        if es_gestionada:
                            viene_en_excel = any(nm.upper() in nombre_bd for nm in nombres_licencias_excel)
                            if not viene_en_excel or not tiene_lic:
                                licencia = asig.licencia
                                asig.estado = 'LIBERADA'
                                asig.activo = False
                                asig.fecha_retiro = timezone.now()
                                asig.save()
                                licencia.estado_operativo = Licencia.ESTADO_VENCIDA if licencia.esta_vencida else Licencia.ESTADO_DISPONIBLE
                                licencia.save(update_fields=['estado_operativo'])
                                liberadas += 1

                    # Aprovisionamiento dinámico de activos
                    texto_extra = f"Complementos M365: {', '.join(skus_extra)}" if skus_extra else ""

                    if tiene_lic:
                        for sku_principal in skus_principales:
                            nombre_lic_bd = mapa_licencias[sku_principal]
                            
                            tiene_esta_licencia = Asignacion.objects.filter(
                                empleado=empleado, 
                                licencia__tipo__nombre__icontains=nombre_lic_bd, 
                                activo=True
                            ).exists()

                            if not tiene_esta_licencia:
                                licencia_libre = Licencia.objects.filter(
                                    tipo__nombre__icontains=nombre_lic_bd,
                                    empresa=empresa_obj,
                                    estado_operativo=Licencia.ESTADO_DISPONIBLE,
                                    fecha_vencimiento__gte=timezone.now().date(),
                                ).exclude(asignaciones__activo=True).first()

                                if licencia_libre:
                                    Asignacion.objects.create(
                                        licencia=licencia_libre,
                                        empleado=empleado,
                                        observaciones=texto_extra
                                    )
                                    licencia_libre.estado_operativo = Licencia.ESTADO_ASIGNADA
                                    licencia_libre.save(update_fields=['estado_operativo'])
                                    asignadas += 1
                                else:
                                    sin_stock += 1 
                            else:
                                # Sincronización de adiciones de bajo nivel (Add-ons)
                                asig_actual = Asignacion.objects.filter(
                                    empleado=empleado, 
                                    licencia__tipo__nombre__icontains=nombre_lic_bd, 
                                    activo=True
                                ).first()
                                if asig_actual and skus_extra:
                                    asig_actual.observaciones = texto_extra
                                    asig_actual.save(update_fields=['observaciones'])

            # ==========================================
            # DIAGNÓSTICO FINAL DE SINCRONIZACIÓN
            # ==========================================
            msg = f"Conciliación Finalizada: {creados} altas, {actualizados} validadas. {asignadas} asignaciones generadas, {liberadas} revocadas."
            if sin_stock > 0:
                msg += f" RIESGO DETECTADO: Déficit de inventario. Faltan {sin_stock} activos en pool corporativo."
                messages.warning(request, msg)
            else:
                messages.success(request, msg)


            log_sincronizar_m365(
                request,
                resumen=f"{creados} altas, {actualizados} validadas, {asignadas} asignaciones, {liberadas} revocadas, sin stock={sin_stock}.",
            )

        except Exception as e:
            messages.error(request, f"Error critico en Pipeline de Datos. Traceback: {str(e)}")

    return redirect('gestionar_licencias')


# ==========================================
# BORRADO MASIVO DE LICENCIAS
# ==========================================
@login_required
def eliminar_licencias_masivo(request):
    """Borrado masivo de licencias. Delega a uc_eliminar_licencias_masivo (omite las asignadas)."""
    from .application.use_cases import uc_eliminar_licencias_masivo
    if request.method == 'POST':
        ids_json = request.POST.get('ids_licencias', '[]')
        try:
            ids = json.loads(ids_json)
            if ids:
                eliminadas, omitidas = uc_eliminar_licencias_masivo(request=request, ids=ids)
                if omitidas:
                    messages.warning(
                        request,
                        f"Se eliminaron {eliminadas} licencias. {omitidas} fueron omitidas por tener asignacion activa."
                    )
                else:
                    messages.success(request, f"Limpieza completada! Se eliminaron {eliminadas} licencias permanentemente.")
            else:
                messages.warning(request, "No se selecciono ninguna licencia para borrar.")
        except Exception as e:
            messages.error(request, f"Ocurrio un error al intentar borrar las licencias: {str(e)}")

    return redirect('gestionar_licencias')


# ==========================================
# CATALOGO DE LICENCIAS (Proveedor + TipoLicencia)
# ==========================================
# Pantalla simple que reemplaza el viejo panel /configuracion/ para las
# entidades que NO migraron a gestion_global (no son CU del CICLO 2).

@login_required
def _texto_importacion(valor):
    if valor is None:
        return ''
    try:
        if pd.isna(valor):
            return ''
    except TypeError:
        pass
    return str(valor).strip()


def _bool_importacion(valor, default=True):
    texto = _texto_importacion(valor).lower()
    if not texto:
        return default
    return texto not in {'0', 'no', 'false', 'falso', 'inactivo', 'inactiva'}


def _valor_fila(fila, columnas):
    for columna in columnas:
        if columna in fila:
            valor = _texto_importacion(fila.get(columna))
            if valor:
                return valor
    return ''


def _importar_proveedores_archivo(request, archivo):
    nombre_archivo = (archivo.name or '').lower()
    if nombre_archivo.endswith('.csv'):
        df = pd.read_csv(archivo)
    else:
        df = pd.read_excel(archivo)

    df = df.rename(columns=lambda col: str(col).strip().lower().replace(' ', '_'))
    creados = 0
    actualizados = 0
    omitidos = 0

    for _, fila in df.iterrows():
        data = fila.to_dict()
        nombre = _valor_fila(data, ['nombre', 'proveedor', 'nombre_comercial'])
        if not nombre:
            omitidos += 1
            continue

        proveedor, creado = Proveedor.objects.get_or_create(nombre=nombre)
        proveedor.razon_social = _valor_fila(data, ['razon_social', 'razonsocial']) or proveedor.razon_social
        proveedor.nit = _valor_fila(data, ['nit', 'ruc', 'identificacion']) or proveedor.nit
        proveedor.contacto = _valor_fila(data, ['contacto', 'persona_contacto', 'persona_de_contacto']) or proveedor.contacto
        proveedor.email = _valor_fila(data, ['email', 'correo', 'correo_electronico']) or proveedor.email
        proveedor.telefono = _valor_fila(data, ['telefono', 'celular', 'phone']) or proveedor.telefono
        proveedor.direccion = _valor_fila(data, ['direccion', 'domicilio']) or proveedor.direccion
        proveedor.sitio_web = _valor_fila(data, ['sitio_web', 'web', 'website']) or proveedor.sitio_web
        proveedor.observaciones = _valor_fila(data, ['observaciones', 'nota', 'notas']) or proveedor.observaciones
        proveedor.activo = _bool_importacion(data.get('activo'), proveedor.activo)
        proveedor.save()

        if creado:
            creados += 1
            log_proveedor_crear(request, proveedor)
        else:
            actualizados += 1
            log_proveedor_editar(request, proveedor)

    return creados, actualizados, omitidos


@login_required
def catalogo_licencias(request):
    exigir_algun_permiso(request, [
        'licencias.view_proveedor',
    ])

    puede_crear_proveedor = request.user.is_superuser or request.user.has_perm('licencias.add_proveedor')
    puede_importar = puede_crear_proveedor

    modal_state = request.session.pop('_lic_catalog_modal_state', None)
    proveedor_form = ProveedorForm(prefix='proveedor') if puede_crear_proveedor else None
    modal_abierto = ''
    q = request.GET.get('q', '').strip()
    estado = request.GET.get('estado', 'activos')

    if request.method == 'POST':
        form_type = request.POST.get('form_type')

        if form_type == 'proveedor':
            if not puede_crear_proveedor:
                messages.error(request, 'No tienes permiso para crear proveedores.')
                return redirect('catalogo_licencias')
            proveedor_form = ProveedorForm(request.POST, prefix='proveedor')
            if proveedor_form.is_valid():
                proveedor = proveedor_form.save()
                log_proveedor_crear(request, proveedor)
                messages.success(request, 'Proveedor creado correctamente.')
                return redirect('catalogo_licencias')
            request.session['_lic_catalog_modal_state'] = {
                'modal': 'proveedor',
                'data': dict(request.POST.items()),
            }
            request.session.modified = True
            return redirect('catalogo_licencias')

        elif form_type == 'importar_proveedores':
            if not puede_importar:
                messages.error(request, 'No tienes permiso para importar proveedores.')
                return redirect('catalogo_licencias')
            archivo = request.FILES.get('archivo_proveedores')
            if not archivo:
                messages.warning(request, 'Selecciona un archivo CSV o Excel.')
                return redirect('catalogo_licencias')
            try:
                creados, actualizados, omitidos = _importar_proveedores_archivo(request, archivo)
                messages.success(
                    request,
                    f'Importacion finalizada: {creados} creados, {actualizados} actualizados, {omitidos} omitidos.'
                )
            except Exception as exc:
                messages.error(request, f'No se pudo importar proveedores: {exc}')
            return redirect('catalogo_licencias')

    elif modal_state:
        modal_abierto = modal_state.get('modal', '')
        data = modal_state.get('data') or None
        if modal_abierto == 'proveedor' and puede_crear_proveedor:
            proveedor_form = ProveedorForm(data, prefix='proveedor')
            proveedor_form.is_valid()
        else:
            modal_abierto = ''

    proveedores = Proveedor.objects.all().order_by('nombre')
    if q:
        proveedores = proveedores.filter(
            Q(nombre__icontains=q)
            | Q(razon_social__icontains=q)
            | Q(nit__icontains=q)
            | Q(contacto__icontains=q)
            | Q(email__icontains=q)
        )
    if estado == 'activos':
        proveedores = proveedores.filter(activo=True)
    elif estado == 'inactivos':
        proveedores = proveedores.filter(activo=False)

    return render(request, 'licencias/catalogo_licencias.html', {
        'titulo': 'Gestionar Proveedores',
        'proveedores': proveedores,
        'kpi_proveedores': {
            'total': Proveedor.objects.count(),
            'activos': Proveedor.objects.filter(activo=True).count(),
            'inactivos': Proveedor.objects.filter(activo=False).count(),
            'sin_correo': Proveedor.objects.filter(email='').count(),
        },
        'proveedor_form': proveedor_form,
        'modal_abierto': modal_abierto,
        'puede_crear_proveedor': puede_crear_proveedor,
        'puede_importar': puede_importar,
        'q': q,
        'estado': estado,
        'licencias_active': 'proveedores',
    })


# ==========================================
# HUB DE LICENCIAS: TABS INTERNOS
# ==========================================
# Pantallas internas del modulo Licencias accesibles desde el navbar del hub:
#   - /licencias/asignaciones/  (todas las asignaciones del sistema)
#   - /licencias/panel/         (dashboard de KPIs)



@login_required
def licencias_dashboard(request):
    """Dashboard ejecutivo con KPIs visuales del modulo de licencias."""
    exigir_algun_permiso(request, ['licencias.view_licencia'])
    from django.utils import timezone
    from datetime import timedelta

    hoy = timezone.now().date()
    proximos_30 = hoy + timedelta(days=30)

    licencias_qs = Licencia.objects.all()
    total_licencias = licencias_qs.count()
    licencias_vencidas = licencias_qs.filter(fecha_vencimiento__lt=hoy).count()
    proximas_vencer = licencias_qs.filter(
        fecha_vencimiento__gte=hoy, fecha_vencimiento__lte=proximos_30
    ).count()
    asignaciones_activas = Asignacion.objects.filter(activo=True).count()
    licencias_disponibles = max(total_licencias - asignaciones_activas, 0)

    # Distribucion por TipoLicencia (top 5)
    from django.db.models import Count
    top_tipos = (
        Licencia.objects.values('tipo__nombre', 'tipo__fabricante')
        .annotate(total=Count('id'))
        .order_by('-total')[:5]
    )

    # Cotizaciones / facturas (si facturacion esta instalada)
    cotizaciones_aprobadas = 0
    facturas_emitidas = 0
    try:
        from facturacion.infrastructure.models import PropuestaLicencia, Factura
        cotizaciones_aprobadas = PropuestaLicencia.objects.filter(estado='APROBADA').count()
        facturas_emitidas = Factura.objects.filter(estado='EMITIDA').count()
    except Exception:
        pass

    return render(request, 'licencias/dashboard_kpis.html', {
        'titulo': 'Dashboard de Licencias',
        'licencias_active': 'dashboard',
        'kpi': {
            'total_licencias': total_licencias,
            'asignaciones_activas': asignaciones_activas,
            'licencias_disponibles': licencias_disponibles,
            'licencias_vencidas': licencias_vencidas,
            'proximas_vencer': proximas_vencer,
            'cotizaciones_aprobadas': cotizaciones_aprobadas,
            'facturas_emitidas': facturas_emitidas,
        },
        'top_tipos': list(top_tipos),
    })


# ==========================================
# AJAX para wizard de asignacion
# ==========================================

@login_required
def ajax_licencias_disponibles(request):
    """
    GET /licencias/ajax/licencias-disponibles/?empresa_id=X

    Retorna licencias de la empresa que NO tienen asignacion activa
    y no estan vencidas.
    """
    from django.utils import timezone
    empresa_id = request.GET.get('empresa_id')
    if not empresa_id:
        return JsonResponse([], safe=False)

    hoy = timezone.now().date()
    qs = (
        Licencia.objects
        .filter(empresa_id=empresa_id, fecha_vencimiento__gte=hoy)
        .exclude(asignaciones__activo=True)
        .select_related('tipo')
        .order_by('tipo__nombre')
    )
    data = [
        {
            'id': l.id,
            'label': f"{l.tipo.nombre} ({l.tipo.fabricante}) - venc {l.fecha_vencimiento}",
            'tipo': l.tipo.nombre,
        }
        for l in qs[:200]
    ]
    return JsonResponse(data, safe=False)


@login_required
def ajax_empleados_empresa(request):
    """GET /licencias/ajax/empleados-empresa/?empresa_id=X"""
    empresa_id = request.GET.get('empresa_id')
    if not empresa_id:
        return JsonResponse([], safe=False)
    qs = Empleado.objects.filter(empresa_id=empresa_id, activo=True).order_by('nombre_completo')
    data = [
        {'id': e.id, 'nombre': e.nombre_completo, 'email': getattr(e, 'email_principal', '')}
        for e in qs[:500]
    ]
    return JsonResponse(data, safe=False)


# ==========================================
# Wizard: Nueva asignacion
# ==========================================

@login_required
def nueva_asignacion(request):
    """Pantalla con wizard Tenant -> Empresa -> Licencia disponible -> Empleado."""
    exigir_permiso(request, 'licencias.add_asignacion')

    if request.method == 'POST':
        from .application.use_cases import uc_asignar_licencia
        from .infrastructure import repositories as repo

        licencia_id = request.POST.get('licencia_id')
        empleado_id = request.POST.get('empleado_id')

        if not licencia_id or not empleado_id:
            messages.error(request, "Debes seleccionar licencia y empleado.")
            return redirect('nueva_asignacion')

        licencia = repo.get_licencia(licencia_id)
        empleado = get_object_or_404(Empleado, pk=empleado_id)

        ok, info = uc_asignar_licencia(request=request, licencia=licencia, empleado=empleado)
        if ok:
            messages.success(request, info)
            return redirect('asignaciones_licencias')
        messages.error(request, info)

    return render(request, 'licencias/nueva_asignacion.html', {
        'titulo': 'Nueva asignacion de licencia',
        'tenants': Tenant.objects.filter(activo=True).order_by('nombre'),
        'licencias_active': 'asignar',
    })


@login_required
def asignar_licencias_masivo(request):
    """Asigna licencias disponibles de un producto a varios empleados de una empresa."""
    from .application.use_cases import uc_asignar_licencia
    exigir_permiso(request, 'licencias.add_asignacion')

    if request.method != 'POST':
        tenant_id = request.GET.get('tenant') or ''
        empresa_id = request.GET.get('empresa') or ''
        empresas = Empresa.objects.filter(activo=True).select_related('tenant').order_by('tenant__nombre', 'nombre')
        if tenant_id:
            empresas = empresas.filter(tenant_id=tenant_id)
        empleados = Empleado.objects.none()
        tipos_disponibles = TipoLicencia.objects.none()
        if empresa_id:
            empleados = Empleado.objects.filter(empresa_id=empresa_id, activo=True).order_by('nombre_completo')
            hoy = timezone.now().date()
            tipos_disponibles = TipoLicencia.objects.filter(
                id__in=Licencia.objects.filter(
                    empresa_id=empresa_id,
                    estado_operativo=Licencia.ESTADO_DISPONIBLE,
                    fecha_vencimiento__gte=hoy,
                ).exclude(asignaciones__activo=True).values('tipo_id')
            ).order_by('fabricante', 'nombre')
        return render(request, 'licencias/asignacion_masiva.html', {
            'tenants': Tenant.objects.filter(activo=True).order_by('nombre'),
            'empresas': empresas,
            'empleados': empleados,
            'tipos_disponibles': tipos_disponibles,
            'tenant_filtro': tenant_id,
            'empresa_filtro': empresa_id,
            'licencias_active': 'asignaciones',
        })

    empresa_id = request.POST.get('empresa_id')
    tipo_id = request.POST.get('tipo_id')
    empleados_ids = request.POST.getlist('empleado_ids')

    if not empresa_id or not tipo_id or not empleados_ids:
        messages.warning(request, "Selecciona empresa, producto y al menos un empleado.")
        return redirect('asignaciones_licencias')

    hoy = timezone.now().date()
    empleados = Empleado.objects.filter(
        id__in=empleados_ids,
        empresa_id=empresa_id,
        activo=True,
    ).select_related('empresa').order_by('nombre_completo')
    licencias_disponibles = list(
        Licencia.objects
        .filter(
            empresa_id=empresa_id,
            tipo_id=tipo_id,
            estado_operativo=Licencia.ESTADO_DISPONIBLE,
            fecha_vencimiento__gte=hoy,
        )
        .exclude(asignaciones__activo=True)
        .select_related('tipo', 'empresa', 'tenant', 'factura_origen')
        .order_by('-factura_origen__fecha', 'fecha_vencimiento', 'id')[:empleados.count()]
    )

    asignadas = 0
    errores = []
    for empleado, licencia in zip(empleados, licencias_disponibles):
        ok, info = uc_asignar_licencia(request=request, licencia=licencia, empleado=empleado)
        if ok:
            asignadas += 1
        else:
            errores.append(info)

    faltantes = max(empleados.count() - len(licencias_disponibles), 0)
    if asignadas:
        messages.success(request, f"Asignacion masiva completada: {asignadas} licencias asignadas.")
    if faltantes:
        messages.warning(request, f"No habia stock suficiente para {faltantes} empleado(s).")
    if errores:
        messages.warning(request, errores[0])

    return redirect(f"{reverse('asignaciones_licencias')}?empresa={empresa_id}&tipo={tipo_id}")


@login_required
def eliminar_asignacion(request, asignacion_id):
    """Elimina fisicamente una asignacion (libera la licencia si estaba activa)."""
    exigir_permiso(request, 'licencias.delete_asignacion')
    asignacion = get_object_or_404(Asignacion, pk=asignacion_id)
    if request.method == 'POST':
        label = f"{asignacion.empleado.nombre_completo} - {asignacion.licencia.tipo.nombre}"
        # Si esta activa, primero liberar para log de bitacora
        if asignacion.activo:
            asignacion.activo = False
            if not asignacion.fecha_retiro:
                asignacion.fecha_retiro = timezone.now()
            asignacion.save()
        asignacion.delete()
        messages.success(request, f"Asignacion eliminada: {label}.")
    return redirect('asignaciones_licencias')


@login_required
def liberar_asignacion(request, asignacion_id):
    """Libera una asignacion (la marca como inactiva). Mantiene historial."""
    exigir_permiso(request, 'licencias.change_asignacion')
    asignacion = get_object_or_404(Asignacion, pk=asignacion_id)
    if request.method == 'POST':
        if asignacion.activo:
            asignacion.activo = False
            if not asignacion.fecha_retiro:
                asignacion.fecha_retiro = timezone.now()
            asignacion.save()
            log_liberar_licencia(request, asignacion.licencia,
                                  empleados=[asignacion.empleado.nombre_completo], cantidad=1)
            messages.success(request, f"Licencia liberada de {asignacion.empleado.nombre_completo}.")
        else:
            messages.warning(request, "Esta asignacion ya estaba liberada.")
    return redirect('asignaciones_licencias')
